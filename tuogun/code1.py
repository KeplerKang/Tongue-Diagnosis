# 代码一：
from typing import Any

from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl

wb1 = load_workbook('表1.xlsx',data_only=True)
wb2 = load_workbook('表2.xlsx',data_only=True)
wb3 = load_workbook('表3.xlsx',data_only=True)
wb4 = load_workbook('表4.xlsx',data_only=True)
# 1.将表2的“组网竣工表”sheet中，表头以及“是否注册”为FALSE的数据复制到表1的“组网竣工表”sheet中，选择性粘贴为文本
# 2.将表2的“宽带电视注册明细”sheet中，表头以及“是否注册”为FALSE的数据复制到表1的“宽带电视注册明细”sheet中，选择性粘贴为文本
# 3.将表2的“和目”sheet中，表头以及“是否注册”为FALSE的数据复制到表1的“和目”sheet中，选择性粘贴为文本
# 4.将表2的“平安乡村”sheet中，“是否注册”为FALSE的数据复制到表1的“平安乡村”sheet中，选择性粘贴为文本
titles = ["组网竣工表","宽带电视注册明细","和目","平安乡村"]
for title in titles:
    sheet1=wb1[title]
    sheet2=wb2[title]
    datas = []
    head = '是否注册'
    cols = sheet2.max_column
    # for i in range(cols):
    #     sheet1.cell(1, i + 1).value = sheet2.cell(1, i + 1).value
    for col in sheet2.columns:
        if col[0].value == head:
            index = col[0].column - 1
    for row in sheet2.rows:
        if row[index].value == False:
            data = []
            for i in range(cols):
                data.append(row[i].value)
            datas.append(data)
    for t in datas:
        sheet1.append(t)
    wb1.save('表1.xlsx')

# 5.将表3的“竣工数据”sheet中，“终端激活（终端与工单出库串码一致）”为否的数据复制到表1的“宽带竣工数据”sheet中，选择性粘贴为文本。
# 在表的第一列插入一列，第一列第一行写装维人员，表4有装维人员和工单编码的对应关系，通过表4查找相应工单编码对应装维人员更新到表1的第一列装维人员中。
sheet5=wb1["宽带竣工数据"]
sheet3=wb3["竣工数据"]
sheet4=wb4['Sheet1']
datas = []
sheet5.cell(1,1).value = '施工人员'
for i in range(sheet3.max_column):
    sheet5.cell(1,i+2).value = sheet3.cell(1, i+1).value

for row in sheet3.rows:
    if row[2].value == '否':
        data = []
        for row1 in sheet4.rows:
            if row1[0].value == row[4].value:
                data.append(row1[24].value)
        for i in range(11):
            data.append(row[i].value)
        data.append(row[23].value)
        data.append(row[24].value)
        datas.append(data)

for t in datas:
    sheet5.append(t)
wb1.save('表1.xlsx')

# 6.将表3的“统计数据”sheet的1-13行复制到表1的“激活率”sheet的39-51行，选择性粘贴为文本
sheet6=wb3["统计数据"]
sheet7=wb1["激活率"]
for i in range(1,14):
    for j in range(1,20):
        if type(sheet6.cell(i,j)).__name__ != 'MergedCell':
            sheet7.cell(i+38,j).value=sheet6.cell(i,j).value
wb1.save('表1.xlsx')

# 代码二：
wb5 = load_workbook('表5.xlsx',data_only=True)
wb6 = load_workbook('表6.xlsx',data_only=True)
wb7 = load_workbook('表7.xlsx',data_only=True)
# 1.取消表5的表头筛选，将表5的B列设备SN和C列的设备'cmei'中的文本类型的数字转换为数字类型
# excel打开时默认显示全部信息，是否删除筛选不影响查看数据，并且openpyxl目前对筛选功能只有添加，还没有清除已有的筛选
sheet5 = wb5['Sheet0']
for i in range(2,sheet5.max_row+1):
    if str(sheet5.cell(i, 2).value).isdigit():
        sheet5.cell(i, 2).value = int(sheet5.cell(i, 2).value)
wb5.save('表5.xlsx')

# 2.对表6的R列第2行至最后一行，如果表6的Q列的第n行在表5的C列能找到，则在表6的R列标记为是，否则标记为否
# 3.对表6的S列第2行至最后一行，如果表6的Q列的第n行在表5的B列能找到，则在表6的S列标记为是，否则标记为否
# 4.表6的T列第n行，如果第R列或第S列为是，则标记为“已激活”，否则标记为未激活
sheet5 = wb5['Sheet0']
sheet6 = wb6['智能组网']
B = []
C = []
for i in range(2,sheet5.max_row+1):
    B.append(str(sheet5.cell(i,2).value))
    C.append(str(sheet5.cell(i,3).value))

for j in range(2,sheet6.max_row+1):
    if str(sheet6.cell(j,17).value) in C:
        sheet6.cell(j,18).value = '是'
    else: sheet6.cell(j,18).value = '否'
    if str(sheet6.cell(j,17).value) in B:
        sheet6.cell(j,19).value = '是'
    else: sheet6.cell(j, 19).value = '否'
for k in range(2,sheet6.max_row+1):
    if sheet6.cell(k,18).value=='是' or sheet6.cell(k,19).value=='是':
        sheet6.cell(k,20).value = '已激活'
    else:
        sheet6.cell(k, 20).value = '未激活'
wb6.save('表6.xlsx')

# 5.表6的S列和R列复制并选择性粘贴同样的位置为纯文本。 ---------------——————？？？？？？
# 6.将表7的宽带sheet和安防sheet复制进表6的宽带sheet和安防sheet，选择性粘贴为粘贴值和数字格式
for title in ["宽带","安防"]:
    ws1 = wb7[title]
    ws2 = wb6[title]
    col1=ws1.max_column
    datas=[]
    for row in ws1.rows:
        data=[]
        for i in range(col1):
            if str(row[i].value).isdigit():
                data.append(int(row[i].value))
            else:
                data.append(row[i].value)
        datas.append(data)
    for t in datas:
        ws2.append(t)
    wb6.save('表6.xlsx')
